"""
-------------------------------------------------------------------------
This file is part of the MindStudio project.
Copyright (c) 2026 Huawei Technologies Co.,Ltd.

MindStudio is licensed under Mulan PSL v2.
You can use this file according to the terms and conditions of the Mulan PSL v2.
You may obtain a copy of Mulan PSL v2 at:

         http://license.coscl.org.cn/MulanPSL2

THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
See the Mulan PSL v2 for more details.
-------------------------------------------------------------------------
"""

import os
import sqlite3
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from trace_analyze import (
    read_task_summary_from_db,
    read_irq_detail_from_db,
    aggregate_by_comm,
    aggregate_by_pid,
    convert_ns_to_us,
    write_excel,
)


class TestDataReading(unittest.TestCase):
    """测试从 db 读取数据"""

    def setUp(self):
        self.tmp_fd, self.tmp_path = tempfile.mkstemp(suffix='.db')
        conn = sqlite3.connect(self.tmp_path)
        cur = conn.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS thread (track_id INTEGER PRIMARY KEY, tid TEXT)")
        cur.execute("CREATE TABLE IF NOT EXISTS slice (track_id INTEGER, name TEXT, duration INTEGER, args TEXT)")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS trace_task_summary (
                comm TEXT, pid INTEGER, cpu_id INTEGER,
                running_ns INTEGER DEFAULT 0, sleeping_ns INTEGER DEFAULT 0,
                runnable_ns INTEGER DEFAULT 0,
                cs_count INTEGER DEFAULT 0, cs_involuntary_count INTEGER DEFAULT 0,
                PRIMARY KEY (comm, pid, cpu_id)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS trace_irq_detail (
                comm TEXT, pid INTEGER, cpu_id INTEGER,
                irq_type TEXT, irq_name TEXT,
                count INTEGER DEFAULT 0, time_ns INTEGER DEFAULT 0,
                PRIMARY KEY (comm, pid, cpu_id, irq_type, irq_name)
            )
        """)
        # 插入测试数据
        cur.executemany(
            "INSERT INTO trace_task_summary VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            [
                ("procA", 100, 0, 1000000, 500000, 200000, 10, 3),
                ("procA", 100, 1, 800000, 300000, 100000, 5, 1),
                ("procA", 100, 2, 600000, 200000, 50000, 3, 0),
                ("procB", 200, 0, 500000, 100000, 30000, 2, 1),
                ("procB", 200, None, 200000, 50000, 10000, 1, 0),
            ],
        )
        cur.executemany(
            "INSERT INTO trace_irq_detail VALUES (?, ?, ?, ?, ?, ?, ?)",
            [
                ("procA", 100, 0, "irq", "timer", 5, 5000),
                ("procA", 100, 1, "irq", "timer", 3, 3000),
                ("procA", 100, 0, "softirq", "net_rx", 2, 2000),
                ("procB", 200, 0, "irq", "rtc", 1, 1000),
            ],
        )
        conn.commit()
        self.conn = conn

    def tearDown(self):
        self.conn.close()
        os.close(self.tmp_fd)
        if os.path.exists(self.tmp_path):
            os.unlink(self.tmp_path)

    def test_read_task_summary(self):
        """测试读取 trace_task_summary 表"""
        df = read_task_summary_from_db(self.conn)
        self.assertEqual(len(df), 5)
        # 第一行
        self.assertEqual(df[0]["comm"], "procA")
        self.assertEqual(df[0]["pid"], 100)
        self.assertEqual(df[0]["cpu_id"], 0)
        self.assertEqual(df[0]["running_ns"], 1000000)

    def test_read_irq_detail(self):
        """测试读取 trace_irq_detail 表"""
        df = read_irq_detail_from_db(self.conn)
        self.assertEqual(len(df), 4)
        self.assertEqual(df[0]["irq_type"], "irq")
        self.assertEqual(df[0]["irq_name"], "timer")
        self.assertEqual(df[0]["count"], 5)


class TestDataAggregation(unittest.TestCase):
    """测试数据聚合逻辑"""

    def test_aggregate_by_comm(self):
        """测试按 comm 汇总（所有 pid 和 cpu 合并）"""
        rows = [
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 0,
                "running_ns": 1000000,
                "sleeping_ns": 500000,
                "runnable_ns": 200000,
                "cs_count": 10,
                "cs_involuntary_count": 3,
            },
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 1,
                "running_ns": 800000,
                "sleeping_ns": 300000,
                "runnable_ns": 100000,
                "cs_count": 5,
                "cs_involuntary_count": 1,
            },
            {
                "comm": "procA",
                "pid": 200,
                "cpu_id": 0,
                "running_ns": 300000,
                "sleeping_ns": 100000,
                "runnable_ns": 50000,
                "cs_count": 2,
                "cs_involuntary_count": 0,
            },
            {
                "comm": "procB",
                "pid": 300,
                "cpu_id": 0,
                "running_ns": 500000,
                "sleeping_ns": 200000,
                "runnable_ns": 80000,
                "cs_count": 4,
                "cs_involuntary_count": 2,
            },
        ]
        result = aggregate_by_comm(rows)
        self.assertEqual(len(result), 2)  # procA, procB

        procA = [r for r in result if r["comm"] == "procA"][0]
        # procA 三个条目合并
        self.assertEqual(procA["running_ns"], 1000000 + 800000 + 300000)
        self.assertEqual(procA["sleeping_ns"], 500000 + 300000 + 100000)
        self.assertEqual(procA["runnable_ns"], 200000 + 100000 + 50000)
        self.assertEqual(procA["cs_count"], 10 + 5 + 2)
        self.assertEqual(procA["cs_involuntary_count"], 3 + 1 + 0)

    def test_aggregate_by_pid(self):
        """测试按 comm:pid 汇总（合并所有 cpu，但 pid 不合并）"""
        rows = [
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 0,
                "running_ns": 1000000,
                "sleeping_ns": 500000,
                "runnable_ns": 200000,
                "cs_count": 10,
                "cs_involuntary_count": 3,
            },
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 1,
                "running_ns": 800000,
                "sleeping_ns": 300000,
                "runnable_ns": 100000,
                "cs_count": 5,
                "cs_involuntary_count": 1,
            },
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 2,
                "running_ns": 600000,
                "sleeping_ns": 200000,
                "runnable_ns": 50000,
                "cs_count": 3,
                "cs_involuntary_count": 0,
            },
            {
                "comm": "procA",
                "pid": 200,
                "cpu_id": 0,
                "running_ns": 300000,
                "sleeping_ns": 100000,
                "runnable_ns": 50000,
                "cs_count": 2,
                "cs_involuntary_count": 0,
            },
        ]
        result = aggregate_by_pid(rows)
        self.assertEqual(len(result), 2)  # procA:100, procA:200

        procA_100 = [r for r in result if r["comm"] == "procA" and r["pid"] == 100][0]
        # procA:100 三个 cpu 合并
        self.assertEqual(procA_100["running_ns"], 1000000 + 800000 + 600000)
        self.assertEqual(procA_100["cs_count"], 10 + 5 + 3)

        procA_200 = [r for r in result if r["comm"] == "procA" and r["pid"] == 200][0]
        self.assertEqual(procA_200["running_ns"], 300000)


class TestUnitConversion(unittest.TestCase):
    """测试 ns → μs 转换"""

    def test_convert_ns_to_us_task(self):
        """测试 task summary 的 ns → μs 转换"""
        rows = [
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 0,
                "running_ns": 1000000,
                "sleeping_ns": 500000,
                "runnable_ns": 200000,
                "cs_count": 10,
                "cs_involuntary_count": 3,
            },
        ]
        result = convert_ns_to_us(rows)
        self.assertEqual(result[0]["running_us"], 1000)
        self.assertEqual(result[0]["sleeping_us"], 500)
        self.assertEqual(result[0]["runnable_us"], 200)
        # 原始 ns 字段应被移除
        self.assertNotIn("running_ns", result[0])

    def test_convert_ns_to_us_irq(self):
        """测试 irq detail 的 ns → μs 转换"""
        rows = [
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 0,
                "irq_type": "irq",
                "irq_name": "timer",
                "count": 5,
                "time_ns": 5000,
            },
        ]
        result = convert_ns_to_us(rows, is_task=False)
        self.assertEqual(result[0]["time_us"], 5)
        self.assertNotIn("time_ns", result[0])


class TestExcelOutput(unittest.TestCase):
    """测试 Excel 输出"""

    def setUp(self):
        self.tmp_fd, self.output_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(self.tmp_fd)

    def tearDown(self):
        if os.path.exists(self.output_path):
            os.unlink(self.output_path)

    def _get_test_data(self):
        task_rows = [
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 0,
                "running_ns": 1000000,
                "sleeping_ns": 500000,
                "runnable_ns": 200000,
                "cs_count": 10,
                "cs_involuntary_count": 3,
            },
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 1,
                "running_ns": 800000,
                "sleeping_ns": 300000,
                "runnable_ns": 100000,
                "cs_count": 5,
                "cs_involuntary_count": 1,
            },
            {
                "comm": "procB",
                "pid": 200,
                "cpu_id": 0,
                "running_ns": 500000,
                "sleeping_ns": 100000,
                "runnable_ns": 30000,
                "cs_count": 2,
                "cs_involuntary_count": 1,
            },
        ]
        irq_rows = [
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 0,
                "irq_type": "irq",
                "irq_name": "timer",
                "count": 5,
                "time_ns": 5000,
            },
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 0,
                "irq_type": "softirq",
                "irq_name": "net_rx",
                "count": 2,
                "time_ns": 2000,
            },
        ]
        return task_rows, irq_rows

    def test_excel_sheet_order(self):
        """测试工作表顺序：by_comm → by_pid → task_summary → irq_detail"""
        task_rows, irq_rows = self._get_test_data()
        write_excel(task_rows, irq_rows, self.output_path)

        from openpyxl import load_workbook

        wb = load_workbook(self.output_path)
        self.assertEqual(wb.sheetnames[0], "task_summary_by_comm")
        self.assertEqual(wb.sheetnames[1], "task_summary_by_pid")
        self.assertEqual(wb.sheetnames[2], "task_summary")
        self.assertEqual(wb.sheetnames[3], "proc_irq_detail")

    def test_excel_has_four_sheets(self):
        """测试 Excel 包含 4 个工作表"""
        task_rows, irq_rows = self._get_test_data()
        write_excel(task_rows, irq_rows, self.output_path)

        from openpyxl import load_workbook

        wb = load_workbook(self.output_path)
        self.assertIn("task_summary", wb.sheetnames)
        self.assertIn("task_summary_by_comm", wb.sheetnames)
        self.assertIn("task_summary_by_pid", wb.sheetnames)
        self.assertIn("proc_irq_detail", wb.sheetnames)

    def test_excel_task_summary_data(self):
        """测试 task_summary 工作表数据正确（ns → μs）"""
        task_rows, irq_rows = self._get_test_data()
        write_excel(task_rows, irq_rows, self.output_path)

        from openpyxl import load_workbook

        wb = load_workbook(self.output_path)
        ws = wb["task_summary"]

        # 第一行是标题
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertIn("running_us", headers)
        self.assertIn("sleeping_us", headers)
        self.assertIn("cs_count", headers)

        # 数据行（3 行原始数据）
        self.assertEqual(ws.max_row, 4)  # 1 header + 3 data

    def test_excel_by_comm_aggregation(self):
        """测试 task_summary_by_comm 聚合正确，无 label 列"""
        task_rows, irq_rows = self._get_test_data()
        write_excel(task_rows, irq_rows, self.output_path)

        from openpyxl import load_workbook

        wb = load_workbook(self.output_path)
        ws = wb["task_summary_by_comm"]

        # 2 个 comm（procA, procB）
        self.assertEqual(ws.max_row, 3)  # 1 header + 2 data

        # 不应有 label 列（用 A 列 comm 作为图表 X 轴）
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertNotIn("label", headers)

    def test_excel_by_pid_no_label(self):
        """测试 task_summary_by_pid 无 label 列"""
        task_rows, irq_rows = self._get_test_data()
        write_excel(task_rows, irq_rows, self.output_path)

        from openpyxl import load_workbook

        wb = load_workbook(self.output_path)
        ws = wb["task_summary_by_pid"]

        # 2 个 pid（procA:100, procB:200）
        self.assertEqual(ws.max_row, 3)  # 1 header + 2 data

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        self.assertNotIn("label", headers)

    def test_excel_by_pid_aggregation(self):
        """测试 task_summary_by_pid 聚合正确"""
        task_rows, irq_rows = self._get_test_data()
        write_excel(task_rows, irq_rows, self.output_path)

        from openpyxl import load_workbook

        wb = load_workbook(self.output_path)
        ws = wb["task_summary_by_pid"]

        # 2 个 pid（procA:100, procB:200）
        self.assertEqual(ws.max_row, 3)  # 1 header + 2 data

    def test_excel_charts_exist(self):
        """测试图表对象存在且位置正确"""
        task_rows, irq_rows = self._get_test_data()
        write_excel(task_rows, irq_rows, self.output_path)

        from openpyxl import load_workbook

        wb = load_workbook(self.output_path)

        # task_summary 不应有图表
        ws_task = wb["task_summary"]
        self.assertEqual(len(ws_task._charts), 0, "task_summary 不应有图表")

        # task_summary_by_comm 应有 Running 图表
        ws_comm = wb["task_summary_by_comm"]
        self.assertTrue(len(ws_comm._charts) > 0, "task_summary_by_comm 应有 Running 图表")

        # proc_irq_detail 应有 2 个图表（Time + Count）
        ws_irq = wb["proc_irq_detail"]
        self.assertTrue(len(ws_irq._charts) >= 2, "proc_irq_detail 应有 2 个 IRQ 图表")

    def test_label_column_visible_in_irq_sheet(self):
        """测试 proc_irq_detail 的 label 列可见（不隐藏，图表 X 轴标签可见）"""
        task_rows, irq_rows = self._get_test_data()
        write_excel(task_rows, irq_rows, self.output_path)

        from openpyxl import load_workbook

        wb = load_workbook(self.output_path)

        # task_summary_by_comm 不应有 label 列
        ws_comm = wb["task_summary_by_comm"]
        headers_comm = [ws_comm.cell(row=1, column=c).value for c in range(1, ws_comm.max_column + 1)]
        self.assertNotIn("label", headers_comm, "task_summary_by_comm 不应有 label 列")

        # proc_irq_detail 应有 label 列且不隐藏
        ws_irq = wb["proc_irq_detail"]
        headers_irq = [ws_irq.cell(row=1, column=c).value for c in range(1, ws_irq.max_column + 1)]
        self.assertIn("label", headers_irq, "proc_irq_detail 应有 label 列")
        label_col_letter = headers_irq.index("label") + 1
        letter = ""
        n = label_col_letter
        while n > 0:
            n, r = divmod(n - 1, 26)
            letter = chr(65 + r) + letter
        self.assertFalse(ws_irq.column_dimensions[letter].hidden, "proc_irq_detail label 列不应隐藏")

    def test_excel_freeze_panes(self):
        """测试冻结首行"""
        task_rows, irq_rows = self._get_test_data()
        write_excel(task_rows, irq_rows, self.output_path)

        from openpyxl import load_workbook

        wb = load_workbook(self.output_path)
        ws = wb["task_summary"]
        self.assertEqual(ws.freeze_panes, "A2")


class TestEmptyData(unittest.TestCase):
    """测试空数据场景"""

    def setUp(self):
        self.tmp_fd, self.output_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(self.tmp_fd)

    def tearDown(self):
        if os.path.exists(self.output_path):
            os.unlink(self.output_path)

    def test_empty_task_rows(self):
        """测试空 task 数据"""
        write_excel([], [], self.output_path)

        from openpyxl import load_workbook

        wb = load_workbook(self.output_path)
        self.assertEqual(len(wb.sheetnames), 4)
        # 只有标题行
        self.assertEqual(wb["task_summary"].max_row, 1)

    def test_empty_irq_rows(self):
        """测试空 irq 数据"""
        task_rows = [
            {
                "comm": "procA",
                "pid": 100,
                "cpu_id": 0,
                "running_ns": 1000000,
                "sleeping_ns": 500000,
                "runnable_ns": 200000,
                "cs_count": 10,
                "cs_involuntary_count": 3,
            },
        ]
        write_excel(task_rows, [], self.output_path)

        from openpyxl import load_workbook

        wb = load_workbook(self.output_path)
        ws = wb["proc_irq_detail"]
        self.assertEqual(ws.max_row, 1)  # 只有标题行


if __name__ == '__main__':
    unittest.main()
