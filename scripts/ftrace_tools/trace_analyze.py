"""
-------------------------------------------------------------------------
This file is part of the MindStudio project.
Copyright (c) 2026 Huawei Technologies Co.,Ltd.

MindStudio is licensed under Mulan PSL v2.
You can use this file according to the terms and conditions of the Mulan PSL v2.
You may obtain a copy of Mulan PSL v2 at:

         http://license.coscl.org.cn/MulanPSL2

THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
See the Mulan PSL v2 for more details.
-------------------------------------------------------------------------
"""

import argparse
import logging
import os
import sqlite3
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment

logging.basicConfig(level=logging.INFO, format='[%(asctime)s] [%(levelname)s]:%(message)s')

DB_PATH_DEFAULT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ftrace_data.db")


def parse_tid(tid_str: str) -> Tuple[str, Optional[int]]:
    """解析 thread.tid 格式为 (comm, pid)。

    常见格式：
    - "comm:pid" -> ("comm", pid)
    - "WatchParentPid::2581884" -> ("WatchParentPid:", 2581884)  (comm 中有冒号)
    - "<idle>:0" -> ("<idle>", 0)
    - "CPU 000" -> ("CPU 000", None)  (CPU 线程，无 pid)

    策略：从右边找最后一个冒号，右边部分必须是纯数字。

    防御性：非字符串输入先转为字符串。
    """
    if not isinstance(tid_str, str):
        tid_str = str(tid_str) if tid_str is not None else ''
    if not tid_str:
        return tid_str, None

    last_colon = tid_str.rfind(':')
    if last_colon == -1:
        return tid_str, None

    potential_pid = tid_str[last_colon + 1 :]
    if potential_pid.isdigit():
        return tid_str[:last_colon], int(potential_pid)

    return tid_str, None


def parse_cpu_from_tid(tid_str: str) -> Optional[int]:
    """从 thread.tid 提取 CPU 编号。

    仅当 tid 是 CPU 线程格式时返回整数，如 "CPU 000" -> 0。
    进程 tid（如 "trace-cmd:3718847"）返回 None。
    """
    if not tid_str:
        return None
    if tid_str.startswith("CPU "):
        num_part = tid_str[4:].strip()
        try:
            return int(num_part)
        except ValueError:
            return None
    return None


@dataclass
class IrqInfo:
    """单个 IRQ 的统计信息"""

    count: int = 0
    time_s: float = 0.0


@dataclass
class TaskStats:
    """单个进程/线程在单个 CPU 上的汇聚统计。

    一个进程可能跨多个 CPU，每个 CPU 单独一行。
    db 内部统一存纳秒（ns），Excel 输出时转换为微秒（μs）。
    """

    comm: str
    pid: int
    cpu_id: Optional[int]  # None 表示未关联到具体 CPU
    running_ns: int = 0
    sleeping_ns: int = 0
    runnable_ns: int = 0
    cs_count: int = 0  # 上下文切换总次数
    cs_involuntary_count: int = 0  # 非自愿切换次数（prev_state='R'，抢占导致）
    irqs: Dict[str, Dict] = field(default_factory=dict)  # key: "irq_type:irq_name"

    def add_duration_ns(self, event_name: str, duration_ns: int) -> None:
        """根据事件名累加纳秒级 duration"""
        if event_name == 'Running':
            self.running_ns += duration_ns
        elif event_name == 'Sleeping':
            self.sleeping_ns += duration_ns
        elif event_name == 'Runnable':
            self.runnable_ns += duration_ns

    def add_irq(self, irq_type: str, irq_name: str, duration_ns: int) -> None:
        """累加 IRQ 统计（ns 存储）"""
        key = f"{irq_type}:{irq_name}"
        if key not in self.irqs:
            self.irqs[key] = {"count": 0, "time_ns": 0, "type": irq_type, "name": irq_name}
        self.irqs[key]["count"] += 1
        self.irqs[key]["time_ns"] += duration_ns


def open_db(db_path: str) -> sqlite3.Connection:
    """打开 SQLite db 文件，返回连接。

    Args:
        db_path: db 文件绝对路径

    Returns:
        sqlite3.Connection

    Raises:
        FileNotFoundError: db 文件不存在
        sqlite3.Error: 数据库连接失败
    """
    if not os.path.isfile(db_path):
        raise FileNotFoundError(f"DB file not found: {db_path}")
    try:
        conn = sqlite3.connect(db_path)
    except sqlite3.Error as e:
        raise sqlite3.Error(f"Failed to connect to database: {db_path}") from e
    conn.row_factory = sqlite3.Row
    return conn


def compute_running_sleeping_runnable_stats(
    conn: sqlite3.Connection,
) -> Dict[Tuple[str, int, Optional[int]], TaskStats]:
    """统计 Running/Sleeping/Runnable 时间，按 (comm, pid, cpu_id) 分组。

    从 slice.args.cpu 字段直接读取 cpu_id（trace_convert 写入）。
    如果 args.cpu 缺失，cpu_id 保持为 None，并打印告警日志。

    Returns:
        {(comm, pid, cpu_id): TaskStats} 字典
    """
    cur = conn.cursor()

    cur.execute("""
        SELECT
            t.tid,
            s.name,
            CAST(json_extract(s.args, '$.cpu') AS INTEGER) as cpu_id,
            SUM(s.duration) as total_ns
        FROM slice s
        JOIN thread t ON s.track_id = t.track_id
        WHERE s.name IN ('Running', 'Sleeping', 'Runnable')
        GROUP BY t.tid, s.name, cpu_id
    """)
    rows = cur.fetchall()
    logging.info("Found %d (tid, event_type, cpu) groups", len(rows))

    stats_map: Dict[Tuple[str, int, Optional[int]], TaskStats] = {}

    for tid, event_name, cpu_id, total_ns in rows:
        if total_ns is None or total_ns == 0:
            continue

        comm, pid = parse_tid(tid)
        if pid is None:
            logging.debug("Skipping tid without pid: %s", tid)
            continue

        key = (comm, pid, cpu_id)
        if key not in stats_map:
            stats_map[key] = TaskStats(comm=comm, pid=pid, cpu_id=cpu_id)

        stats_map[key].add_duration_ns(event_name, total_ns)

    # 检查是否有 cpu_id=None 的记录
    no_cpu_entries = sum(1 for k in stats_map if k[2] is None)
    if no_cpu_entries > 0:
        logging.warning(
            "%d entries have no cpu_id — events were generated by an old trace_convert "
            "that lacks cpu in args. CPU column will be blank.",
            no_cpu_entries,
        )

    logging.info("Generated %d TaskStats entries", len(stats_map))
    return stats_map


# 内核进程关键词，与 trace_convert.py 中的 KERNEL_PROCESS 保持一致
_KERNEL_PROCESS_KEYWORDS = ['migration', 'swapper', 'kworker']


def _is_kernel_process(comm: str) -> bool:
    """判断是否为内核进程。

    与 trace_convert.py 中的 is_kernel_process 逻辑保持一致。
    """
    for keyword in _KERNEL_PROCESS_KEYWORDS:
        if keyword in comm:
            return True
    return False


def compute_irq_stats(conn: sqlite3.Connection, stats_map: Dict[Tuple[str, int, Optional[int]], TaskStats]) -> None:
    """统计 IRQ 和 SoftIRQ 打断进程的次数与耗时，写入 TaskStats.irqs。

    从进程视角出发：每条 irq/softirq 事件的 args.task 字段记录了**被中断的进程**。
    按 (task, irq_name/action, cpu_id) 分组聚合后，定位到 stats_map 中的对应行并累加。

    过滤规则：
    - task 字段缺失 → 跳过
    - task 是 idle 进程（<idle>）→ 跳过
    - task 不在 stats_map 中 → 跳过（说明该进程没有 running/sleeping 事件）

    Args:
        conn: SQLite 数据库连接
        stats_map: 由 compute_running_sleeping_runnable_stats() 返回的 TaskStats 字典
    """
    cur = conn.cursor()

    # 合并查询：一次扫描同时获取 irq 和 softirq，减少全表扫描次数
    cur.execute("""
        SELECT
            json_extract(s.args, '$.task') as interrupted_task,
            s.name as event_type,
            CASE
                WHEN s.name = 'irq' THEN json_extract(s.args, '$.name')
                ELSE json_extract(s.args, '$.action')
            END as irq_action,
            CAST(REPLACE(t.tid, 'CPU ', '') AS INTEGER) as cpu_id,
            COUNT(*) as count,
            CAST(SUM(s.duration) AS INTEGER) as total_ns
        FROM slice s
        JOIN thread t ON s.track_id = t.track_id
        WHERE s.name IN ('irq', 'softirq')
          AND t.tid LIKE 'CPU %'
          AND interrupted_task IS NOT NULL
        GROUP BY interrupted_task, event_type, irq_action, cpu_id
    """)
    rows = cur.fetchall()
    irq_count = 0
    softirq_count = 0
    skipped_idle = 0
    skipped_no_match = 0

    for interrupted_task, event_type, irq_action, cpu_id, count, total_ns in rows:
        # 健壮性：irq_action 或 count 无效则跳过
        if irq_action is None or count == 0:
            continue
        if not isinstance(interrupted_task, str):
            logging.debug("Skipping irq event with non-string task: %r", interrupted_task)
            continue

        comm, pid = parse_tid(interrupted_task)
        if comm == '<idle>':
            skipped_idle += count
            continue

        key = (comm, pid, cpu_id)
        if key in stats_map:
            _add_irq_to_entry(stats_map[key], event_type, irq_action, count, total_ns)
            if event_type == 'irq':
                irq_count += count
            else:
                softirq_count += count
        else:
            skipped_no_match += count

    if irq_count > 0 or softirq_count > 0:
        logging.info("Aggregated %d irq events and %d softirq events into TaskStats", irq_count, softirq_count)
    else:
        logging.info("No irq/softirq events found for tracked tasks")

    if skipped_idle > 0:
        logging.debug("Skipped %d idle task irq/softirq events", skipped_idle)
    if skipped_no_match > 0:
        logging.debug("Skipped %d irq events with no matching TaskStats", skipped_no_match)


def _add_irq_to_entry(task_stats: TaskStats, irq_type: str, irq_name: str, count: int, total_ns: int) -> None:
    """将聚合后的 irq 统计累加到单个 TaskStats 条目。

    count 和 total_ns 已由 SQL 聚合完毕，直接设置到 irqs 字典中。
    """
    key = f"{irq_type}:{irq_name}"
    if key not in task_stats.irqs:
        task_stats.irqs[key] = {"count": 0, "time_ns": 0, "type": irq_type, "name": irq_name}
    task_stats.irqs[key]["count"] += count
    task_stats.irqs[key]["time_ns"] += total_ns


def _create_result_tables(conn: sqlite3.Connection) -> None:
    """创建结果表（trace_task_summary 和 trace_irq_detail）。

    使用 CREATE TABLE IF NOT EXISTS，支持重复调用。
    """
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS trace_task_summary (
            comm TEXT,
            pid INTEGER,
            cpu_id INTEGER,
            running_ns INTEGER DEFAULT 0,
            sleeping_ns INTEGER DEFAULT 0,
            runnable_ns INTEGER DEFAULT 0,
            cs_count INTEGER DEFAULT 0,
            cs_involuntary_count INTEGER DEFAULT 0,
            PRIMARY KEY (comm, pid, cpu_id)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS trace_irq_detail (
            comm TEXT,
            pid INTEGER,
            cpu_id INTEGER,
            irq_type TEXT,
            irq_name TEXT,
            count INTEGER DEFAULT 0,
            time_ns INTEGER DEFAULT 0,
            PRIMARY KEY (comm, pid, cpu_id, irq_type, irq_name)
        )
    """)


def _build_task_rows(stats_map: Dict[Tuple[str, int, Optional[int]], TaskStats]) -> list:
    """从 stats_map 构建 trace_task_summary 插入用的元组列表。

    Returns:
        [(comm, pid, cpu_id, running_ns, sleeping_ns, runnable_ns,
          cs_count, cs_involuntary_count), ...]
    """
    rows = []
    for (comm, pid, cpu_id), stats in stats_map.items():
        rows.append(
            (
                comm,
                pid,
                cpu_id,
                stats.running_ns,
                stats.sleeping_ns,
                stats.runnable_ns,
                stats.cs_count,
                stats.cs_involuntary_count,
            )
        )
    return rows


def _build_irq_rows(stats_map: Dict[Tuple[str, int, Optional[int]], TaskStats]) -> list:
    """从 stats_map 构建 trace_irq_detail 插入用的元组列表。

    仅包含有 IRQ 记录的条目。

    Returns:
        [(comm, pid, cpu_id, irq_type, irq_name, count, time_ns), ...]
    """
    rows = []
    for (comm, pid, cpu_id), stats in stats_map.items():
        for irq_key, irq_info in stats.irqs.items():
            rows.append(
                (
                    comm,
                    pid,
                    cpu_id,
                    irq_info["type"],
                    irq_info["name"],
                    irq_info["count"],
                    irq_info["time_ns"],
                )
            )
    return rows


def write_results_to_db(conn: sqlite3.Connection, stats_map: Dict[Tuple[str, int, Optional[int]], TaskStats]) -> None:
    """将 stats_map 统计结果写回 db 新表。

    策略：DELETE 旧数据 + INSERT（显式事务 + executemany 批量插入）。
    性能优化：关闭同步 + 内存 journal（离线分析场景安全）。

    Args:
        conn: SQLite 数据库连接
        stats_map: 由 compute_* 系列函数填充的 TaskStats 字典
    """
    cur = conn.cursor()

    # 创建结果表
    _create_result_tables(conn)

    # 性能优化：关闭同步 + 内存 journal（必须在事务外设置）
    # 离线分析工具场景安全：输入 db 可重新生成，无并发写入
    cur.execute("PRAGMA synchronous = OFF")
    cur.execute("PRAGMA journal_mode = MEMORY")

    # 清空旧数据（避免重复运行产生脏数据）
    cur.execute("DELETE FROM trace_task_summary")
    cur.execute("DELETE FROM trace_irq_detail")

    # 构建数据行
    task_rows = _build_task_rows(stats_map)
    irq_rows = _build_irq_rows(stats_map)

    logging.info("Writing %d task rows and %d irq rows to db", len(task_rows), len(irq_rows))

    # 批量写入（SQLite 默认 autocommit=False，DELETE 已开始隐式事务）
    if task_rows:
        cur.executemany(
            "INSERT INTO trace_task_summary "
            "(comm, pid, cpu_id, running_ns, sleeping_ns, runnable_ns, "
            "cs_count, cs_involuntary_count) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            task_rows,
        )
    if irq_rows:
        cur.executemany(
            "INSERT INTO trace_irq_detail "
            "(comm, pid, cpu_id, irq_type, irq_name, count, time_ns) "
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            irq_rows,
        )
    conn.commit()

    logging.info("Results written to db successfully")


def compute_cs_count(conn: sqlite3.Connection, stats_map: Dict[Tuple[str, int, Optional[int]], TaskStats]) -> None:
    """统计上下文切换次数，按 (comm, pid, cpu_id) 分组。

    从 sched_switch 事件的 args 中提取 prev_comm/prev_pid/prev_state，
    结合 CPU 线程的 tid 获取 cpu_id，定位到 stats_map 中的对应行并累加 cs_count。

    内核进程（swapper/kworker/migration）的 prev 事件被跳过。

    Args:
        conn: SQLite 数据库连接
        stats_map: 由 compute_running_sleeping_runnable_stats() 返回的 TaskStats 字典
    """
    cur = conn.cursor()

    cur.execute("""
        SELECT
            json_extract(s.args, '$.prev_comm') as prev_comm,
            json_extract(s.args, '$.prev_pid') as prev_pid,
            json_extract(s.args, '$.prev_state') as prev_state,
            CAST(REPLACE(t.tid, 'CPU ', '') AS INTEGER) as cpu_id
        FROM slice s
        JOIN thread t ON s.track_id = t.track_id
        WHERE s.name = 'sched_switch'
          AND t.tid LIKE 'CPU %'
    """)
    rows = cur.fetchall()
    logging.info("Found %d sched_switch events on CPU threads", len(rows))

    cs_count = 0
    cs_involuntary_count = 0
    skipped_kernel = 0

    for prev_comm, prev_pid_str, prev_state, cpu_id in rows:
        if prev_comm is None or prev_pid_str is None:
            continue

        # 跳过内核进程
        if _is_kernel_process(prev_comm):
            skipped_kernel += 1
            continue

        try:
            prev_pid = int(prev_pid_str)
        except (ValueError, TypeError):
            continue

        key = (prev_comm, prev_pid, cpu_id)
        if key in stats_map:
            stats_map[key].cs_count += 1
            cs_count += 1

            # prev_state='R' 表示进程仍在 Ready 状态被强制换下（非自愿/抢占）
            if prev_state == 'R':
                stats_map[key].cs_involuntary_count += 1
                cs_involuntary_count += 1

    if skipped_kernel > 0:
        logging.debug("Skipped %d kernel process sched_switch events", skipped_kernel)

    logging.info("Accumulated %d context switch events (%d involuntary) into TaskStats", cs_count, cs_involuntary_count)


def build_arg_parser() -> argparse.ArgumentParser:
    """构建命令行参数解析器"""
    parser = argparse.ArgumentParser(description='Analyze ftrace data from trace_convert db output')
    parser.add_argument(
        '--input',
        '-i',
        default=DB_PATH_DEFAULT,
        help=f'Input db file path from trace_convert.py (default: {DB_PATH_DEFAULT})',
    )
    parser.add_argument('--output', '-o', default=None, help='Output Excel report path (default: <input>_report.xlsx)')
    return parser


def main():
    parser = build_arg_parser()
    args = parser.parse_args()

    db_path = args.input
    output_path = args.output

    if not output_path:
        base = os.path.splitext(db_path)[0]
        output_path = f"{base}_report.xlsx"

    logging.info("Opening db: %s", db_path)
    conn = open_db(db_path)

    try:
        # Step 1: 统计 Running/Sleeping/Runnable 时间
        logging.info("Computing running/sleeping/runnable stats...")
        stats_map = compute_running_sleeping_runnable_stats(conn)

        # Step 2: 统计上下文切换
        logging.info("Computing context switch stats...")
        compute_cs_count(conn, stats_map)

        # Step 3: 统计 IRQ
        logging.info("Computing IRQ stats...")
        compute_irq_stats(conn, stats_map)

        # Step 4: 写回 db
        logging.info("Writing results to db...")
        write_results_to_db(conn, stats_map)

        # Step 5: 生成 Excel 报告
        logging.info("Generating Excel report: %s", output_path)
        task_rows = read_task_summary_from_db(conn)
        irq_rows = read_irq_detail_from_db(conn)
        write_excel(task_rows, irq_rows, output_path)

        logging.info("Analysis complete. Report saved to: %s", output_path)
    finally:
        conn.close()


# ============================================================
# Excel 输出与图表生成
# ============================================================


def read_task_summary_from_db(conn: sqlite3.Connection) -> List[Dict]:
    """从 trace_task_summary 表读取所有数据。

    Returns:
        列表，每个元素是包含所有字段的 dict
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT comm, pid, cpu_id, running_ns, sleeping_ns, runnable_ns,
               cs_count, cs_involuntary_count
        FROM trace_task_summary
        ORDER BY running_ns DESC
    """)
    columns = [d[0] for d in cur.description]
    return [dict(zip(columns, row)) for row in cur.fetchall()]


def read_irq_detail_from_db(conn: sqlite3.Connection) -> List[Dict]:
    """从 trace_irq_detail 表读取所有数据。

    Returns:
        列表，每个元素是包含所有字段的 dict
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT comm, pid, cpu_id, irq_type, irq_name, count, time_ns
        FROM trace_irq_detail
        ORDER BY time_ns DESC
    """)
    columns = [d[0] for d in cur.description]
    return [dict(zip(columns, row)) for row in cur.fetchall()]


def aggregate_by_comm(rows: List[Dict]) -> List[Dict]:
    """按 comm 汇总（所有 pid 和 cpu 合并）。

    Args:
        rows: 原始 task summary 数据（含 comm, pid, cpu_id）

    Returns:
        按 comm 聚合后的列表，只保留 comm + 聚合后的数值字段
    """
    comm_map: Dict[str, Dict] = {}
    for row in rows:
        comm = row["comm"]
        if comm not in comm_map:
            comm_map[comm] = {
                "comm": comm,
                "running_ns": 0,
                "sleeping_ns": 0,
                "runnable_ns": 0,
                "cs_count": 0,
                "cs_involuntary_count": 0,
            }
        entry = comm_map[comm]
        entry["running_ns"] += row.get("running_ns", 0)
        entry["sleeping_ns"] += row.get("sleeping_ns", 0)
        entry["runnable_ns"] += row.get("runnable_ns", 0)
        entry["cs_count"] += row.get("cs_count", 0)
        entry["cs_involuntary_count"] += row.get("cs_involuntary_count", 0)

    return sorted(comm_map.values(), key=lambda x: x["running_ns"], reverse=True)


def aggregate_by_pid(rows: List[Dict]) -> List[Dict]:
    """按 comm:pid 汇总（合并所有 cpu，但 pid 不合并）。

    Args:
        rows: 原始 task summary 数据

    Returns:
        按 (comm, pid) 聚合后的列表
    """
    pid_map: Dict[Tuple[str, int], Dict] = {}
    for row in rows:
        key = (row["comm"], row["pid"])
        if key not in pid_map:
            pid_map[key] = {
                "comm": row["comm"],
                "pid": row["pid"],
                "running_ns": 0,
                "sleeping_ns": 0,
                "runnable_ns": 0,
                "cs_count": 0,
                "cs_involuntary_count": 0,
            }
        entry = pid_map[key]
        entry["running_ns"] += row.get("running_ns", 0)
        entry["sleeping_ns"] += row.get("sleeping_ns", 0)
        entry["runnable_ns"] += row.get("runnable_ns", 0)
        entry["cs_count"] += row.get("cs_count", 0)
        entry["cs_involuntary_count"] += row.get("cs_involuntary_count", 0)

    return sorted(pid_map.values(), key=lambda x: x["running_ns"], reverse=True)


def convert_ns_to_us(rows: List[Dict], is_task: bool = True) -> List[Dict]:
    """将纳秒字段转换为微秒。

    Args:
        rows: 原始数据
        is_task: True 为 task summary（转换 running/sleeping/runnable_ns），
                 False 为 irq detail（转换 time_ns）

    Returns:
        新列表，ns 字段替换为 us 字段
    """
    result = []
    for row in rows:
        new_row = dict(row)
        if is_task:
            new_row["running_us"] = new_row.pop("running_ns", 0) / 1000
            new_row["sleeping_us"] = new_row.pop("sleeping_ns", 0) / 1000
            new_row["runnable_us"] = new_row.pop("runnable_ns", 0) / 1000
        else:
            new_row["time_us"] = new_row.pop("time_ns", 0) / 1000
        result.append(new_row)
    return result


def write_excel(task_rows: List[Dict], irq_rows: List[Dict], output_path: str) -> None:
    """生成 Excel 报告，包含 4 个工作表和 3 个图表。

    工作表（按顺序）：
    - task_summary_by_comm: 按 comm 汇总 + Top10 Running 柱状图
    - task_summary_by_pid: 按 comm:pid 汇总（cpu 合并）
    - task_summary: 每行一个 (comm, pid, cpu_id)，ns → μs
    - proc_irq_detail: IRQ 明细 + Top10 Time/Count 柱状图

    图表位置：数据从 A1 开始，图表放在右侧（I1 起）竖直堆叠。
    """

    wb = Workbook()

    # 准备数据
    task_us = convert_ns_to_us(task_rows, is_task=True)
    irq_us = convert_ns_to_us(irq_rows, is_task=False)
    by_comm = aggregate_by_comm(task_rows)
    by_comm_us = convert_ns_to_us(by_comm, is_task=True)
    # 为 IRQ 图表生成 label（写在 sheet 最后一列）
    for row in irq_us:
        cpu = row.get("cpu_id", "")
        cpu_str = f"@{cpu}" if cpu is not None else ""
        row["label"] = f"{row['comm']}:{row['pid']}{cpu_str}"
    by_pid = aggregate_by_pid(task_rows)
    by_pid_us = convert_ns_to_us(by_pid, is_task=True)

    # --- Sheet 1: task_summary_by_comm (+ Running chart) ---
    ws_comm = wb.active
    ws_comm.title = "task_summary_by_comm"
    _write_sheet_data(ws_comm, by_comm_us, freeze=True)
    if by_comm_us:
        # 直接用 A 列（comm）作为 X 轴标签
        _insert_running_chart(ws_comm, by_comm_us, "I1")

    # --- Sheet 2: task_summary_by_pid ---
    ws_pid = wb.create_sheet("task_summary_by_pid")
    _write_sheet_data(ws_pid, by_pid_us, freeze=True)

    # --- Sheet 3: task_summary ---
    ws_task = wb.create_sheet("task_summary")
    _write_sheet_data(ws_task, task_us, freeze=True)

    # --- Sheet 4: proc_irq_detail (+ IRQ Time/Count charts) ---
    ws_irq = wb.create_sheet("proc_irq_detail")
    _write_sheet_data(ws_irq, irq_us, freeze=True)
    if irq_us:
        _insert_irq_time_chart(ws_irq, irq_us, "I1")
        _insert_irq_count_chart(ws_irq, irq_us, "I22")

    wb.save(output_path)
    logging.info("Excel report saved: %s", output_path)


def _write_sheet_data(ws, rows: List[Dict], exclude_keys: Optional[set] = None, freeze: bool = False) -> None:
    """将数据写入工作表（标题行 + 数据行）。

    Args:
        ws: openpyxl worksheet
        rows: 数据列表（dict）
        exclude_keys: 不写入 sheet 的 key 集合
        freeze: 是否冻结首行
    """
    if not rows:
        return

    # 过滤掉 exclude_keys 中的字段
    filtered = [{k: v for k, v in row.items() if k not in (exclude_keys or set())} for row in rows]

    # 标题行
    headers = list(filtered[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for row_idx, row in enumerate(filtered, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row.get(header)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动列宽
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in filtered:
            val = row.get(header, "")
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 30)

    if freeze:
        ws.freeze_panes = "A2"


def _get_col_letter(rows: List[Dict], key: str) -> Optional[str]:
    """获取某个 key 对应的 Excel 列字母。"""
    if not rows:
        return None
    headers = list(rows[0].keys())
    if key not in headers:
        return None
    col_idx = headers.index(key) + 1
    result = ""
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _insert_running_chart(ws, rows: List[Dict], anchor_cell: str) -> None:
    """插入 Top10 Running 时间柱状图。

    直接使用 A 列（comm）作为 X 轴分类。
    """
    if not rows:
        return

    headers = list(rows[0].keys())
    running_col = headers.index("running_us") + 1

    data_end = min(11, len(rows) + 1)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Top10 Running Time"
    chart.style = 10
    chart.y_axis.title = "Running Time (us)"

    cat_ref = Reference(ws, min_col=1, min_row=2, max_row=data_end)  # A 列 = comm
    val_ref = Reference(ws, min_col=running_col, min_row=1, max_row=data_end)

    chart.add_data(val_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.shape = 4

    ws.add_chart(chart, anchor_cell)


def _insert_irq_time_chart(ws, rows: List[Dict], anchor_cell: str) -> None:
    """插入 Top10 IRQ Time 柱状图。

    使用 label 列作为 X 轴，time_us 列作为值。
    """
    if not rows:
        return

    headers = list(rows[0].keys())
    label_col = headers.index("label") + 1
    time_col = headers.index("time_us") + 1

    data_end = min(11, len(rows) + 1)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Top10 IRQ Time"
    chart.style = 10
    chart.y_axis.title = "IRQ Time (us)"

    cat_ref = Reference(ws, min_col=label_col, min_row=2, max_row=data_end)
    val_ref = Reference(ws, min_col=time_col, min_row=1, max_row=data_end)

    chart.add_data(val_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.shape = 4

    ws.add_chart(chart, anchor_cell)


def _insert_irq_count_chart(ws, rows: List[Dict], anchor_cell: str) -> None:
    """插入 Top10 IRQ Count 柱状图。

    使用 label 列作为 X 轴，count 列作为值。
    """
    if not rows:
        return

    headers = list(rows[0].keys())
    label_col = headers.index("label") + 1
    count_col = headers.index("count") + 1

    data_end = min(11, len(rows) + 1)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Top10 IRQ Count"
    chart.style = 10
    chart.y_axis.title = "IRQ Count"

    cat_ref = Reference(ws, min_col=label_col, min_row=2, max_row=data_end)
    val_ref = Reference(ws, min_col=count_col, min_row=1, max_row=data_end)

    chart.add_data(val_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.shape = 4

    ws.add_chart(chart, anchor_cell)


if __name__ == '__main__':
    main()
